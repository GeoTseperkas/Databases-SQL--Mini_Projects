import cx_Oracle
import openpyxl
import os
from datetime import datetime, timedelta
import tkinter as tk

# Create a Tkinter window to get the start and end dates from the user
root = tk.Tk()
root.title("Task Report Generator")

# Set the background color of all widgets inside the root window to gray
for child in root.winfo_children():
    child.configure(bg='gray')

# Create StringVar objects to store the start and end dates entered by the user
start_date_var = tk.StringVar()
end_date_var = tk.StringVar()

# Define a function to destroy the window when the user submits the dates
def submit_dates():
    root.destroy()

# Create labels and entry fields for the start and end dates
start_date_label = tk.Label(root, text="Start date (DD/MM/YYYY): ")
start_date_label.grid(row=0, column=0)
start_date_entry = tk.Entry(root, textvariable=start_date_var,width=30)
start_date_entry.grid(row=0, column=1)

end_date_label = tk.Label(root, text="End date (DD/MM/YYYY): ")
end_date_label.grid(row=1, column=0)
end_date_entry = tk.Entry(root, textvariable=end_date_var,width=30)
end_date_entry.grid(row=1, column=1)

# Create a submit button to finalize the dates entered by the user
submit_button = tk.Button(root, text="Submit", command=submit_dates)
submit_button.grid(row=2, column=0, columnspan=2)

# Center the window on the screen
window_width = 300
window_height = 100
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = int((screen_width/2) - (window_width/2))
y = int((screen_height/2) - (window_height/2))
root.geometry("{}x{}+{}+{}".format(window_width, window_height, x, y))

# Start the Tkinter mainloop to display the window and wait for user input
root.mainloop()

# Retrieve the start and end dates entered by the user from the StringVar objects
start_date_str = start_date_var.get()
end_date_str = end_date_var.get()

# Convert the start and end date strings to datetime objects for use in SQL queries
start_date = datetime.strptime(start_date_str, '%d/%m/%Y')
end_date = datetime.strptime(end_date_str, '%d/%m/%Y')

# Set up the connection string with placeholders for the username, password, host, port, and service name
connection_string = '{username}/{password}@{host}:{port}/{service_name}'

# Fill in the placeholders with the actual values for your database
username = 'your_username'
password = 'your_password'
host = 'your_host'
port = 'your_port'
service_name = 'your_service_name'

# Create a connection using the filled-in connection string
connection = cx_Oracle.connect(connection_string.format(
    username=username, password=password, host=host, port=port, service_name=service_name
))

# Construct the dynamic SQL statement
sql_query = '''
SELECT name, ''' + 
', '.join(['COALESCE(SUM(CASE WHEN t.date = TO_DATE(\'{}\', \'DD/MM/YYYY\') THEN t.jobs END), 0) 
AS "{}"'.format((start_date + timedelta(days=i)).strftime('%d/%m/%Y'), (start_date + timedelta(days=i))
.strftime('%d/%m/%Y')) for i in range((end_date - start_date).days + 1)]) + ''', COALESCE(SUM(jobs), 0) AS "Grand Total"
FROM (
    SELECT name, date, jobs FROM Sales
    UNION ALL
    SELECT name, date, jobs FROM Imports
    UNION ALL
    SELECT name, date, jobs FROM Revenue_Analysis
    UNION ALL
    SELECT name, date, jobs FROM Profit_Analysis
) t
WHERE t.date >= TO_DATE(\'{}\', \'DD/MM/YYYY\') AND t.INS_DATE <= TO_DATE(\'{}\', \'DD/MM/YYYY\')
GROUP BY name
ORDER BY name'''.format(start_date_str, end_date_str)

# Construct secondary SQL statement
sql_query1 = '''
SELECT
  n.name,
  COALESCE(t1.jobs, 0) AS Sales,
  COALESCE(t2.jobs, 0) AS Imports,
  COALESCE(t3.jobs, 0) AS Revenue_Analysis,
  COALESCE(t4.jobs, 0) AS Profit_Analysis,
  COALESCE(t1.jobs, 0) + COALESCE(t2.jobs, 0) + COALESCE(t3.jobs, 0) + COALESCE(t4.jobs, 0) AS Grand_Total
FROM
  (SELECT name, date, jobs FROM Sales
   UNION ALL
   SELECT name, date, jobs FROM Imports
   UNION ALL
   SELECT name, date, jobs FROM Revenue_Analysis
   UNION ALL
   SELECT name, date, jobs FROM Profit_Analysis) n
LEFT JOIN
  (SELECT name, SUM(jobs) AS jobs FROM Sales 
  WHERE date BETWEEN TO_DATE('{}', 'DD/MM/YYYY') AND TO_DATE('{}', 'DD/MM/YYYY') GROUP BY name) t1
ON n.name = t1.name
LEFT JOIN
  (SELECT name, SUM(jobs) AS jobs FROM Imports 
  WHERE date BETWEEN TO_DATE('{}', 'DD/MM/YYYY') AND TO_DATE('{}', 'DD/MM/YYYY') GROUP BY name) t2
ON n.name = t2.name
LEFT JOIN
  (SELECT name, SUM(jobs) AS jobs FROM Revenue_Analysis 
  WHERE date BETWEEN TO_DATE('{}', 'DD/MM/YYYY') AND TO_DATE('{}', 'DD/MM/YYYY') GROUP BY name) t3
ON n.name = t3.name
LEFT JOIN
  (SELECT name, SUM(jobs) AS jobs FROM Profit_Analysis 
  WHERE date BETWEEN TO_DATE('{}', 'DD/MM/YYYY') AND TO_DATE('{}', 'DD/MM/YYYY') GROUP BY name) t4
ON n.name = t4.name
WHERE COALESCE(t1.jobs, 0) + COALESCE(t2.jobs, 0) + COALESCE(t3.jobs, 0) + COALESCE(t4.jobs, 0) > 0
ORDER BY n.name
'''.format(start_date.strftime('%d/%m/%Y'), end_date.strftime('%d/%m/%Y'), 
           start_date.strftime('%d/%m/%Y'), end_date.strftime('%d/%m/%Y'), 
           start_date.strftime('%d/%m/%Y'), end_date.strftime('%d/%m/%Y'), 
           start_date.strftime('%d/%m/%Y'), end_date.strftime('%d/%m/%Y'))


# Execute the dynamic SQL statements
cursor = connection.cursor()

# Execute the first query
cursor.execute(sql_query)
# Fetch the results
results = cursor.fetchall()

# Create a new Excel workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Task Report"

# Write the headers to the worksheet
headers = [desc[0] for desc in cursor.description]
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)

# Write the results to the worksheet
for row_num, row in enumerate(results, 2):
    for col_num, cell_value in enumerate(row, 1):
        ws.cell(row=row_num, column=col_num, value=cell_value)


# Execute the second query
cursor.execute(sql_query1)
# Fetch the results
results1 = cursor.fetchall()

# Create a new worksheet for the second query
ws1 = wb.create_sheet(title="Task Analysis")

# Write the headers to the worksheet
headers1 = [desc[0] for desc in cursor.description]
for col_num, header in enumerate(headers1, 1):
    ws1.cell(row=1, column=col_num, value=header)

# Write the results to the worksheet
for row_num, row in enumerate(results1, 2):
    for col_num, cell_value in enumerate(row, 1):
        ws1.cell(row=row_num, column=col_num, value=cell_value)

# Create the folder if it does not exist
if not os.path.exists('Task Report'):
    os.mkdir('Task Report')

# Save the workbook inside the folder
wb.save(os.path.join('Task Report', 'Task Report.xlsx'))

# Close the cursor and connection
cursor.close()
connection.close()